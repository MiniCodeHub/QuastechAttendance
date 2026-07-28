import os
import re
import math
from datetime import datetime, date, time, timedelta
from io import BytesIO
from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file
import pymysql
from pymysql.cursors import DictCursor
import pytz
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(BASE_DIR)
TEMPLATES_DIR = os.path.join(PROJECT_ROOT, 'frontend', 'templates')
STATIC_DIR = os.path.join(PROJECT_ROOT, 'frontend', 'static')

app = Flask(
    __name__,
    template_folder=TEMPLATES_DIR,
    static_folder=STATIC_DIR
)
app.secret_key = os.urandom(24)

# ============================================================
# DATABASE CONFIGURATION – LOCAL MYSQL (XAMPP)
# ============================================================
DB_CONFIG = {
    "host": "gateway01.ap-southeast-1.prod.aws.tidbcloud.com",
    "port": 4000,
    "user": "3dtXqyjkbNdTH2t.root",
    "password": "vzuHZOlyqLj195LO",
    "database": "defaultdb",
    "charset": "utf8mb4",
    "cursorclass": DictCursor,
    "autocommit": True,
    "ssl_verify_cert": True,
    "ssl_verify_identity": True,
    "ssl_ca": os.path.join(BASE_DIR, "isrgrootx1.pem")
}

ADMIN_CODE = 'admin1246'

COLLEGE_LAT = 19.18748
COLLEGE_LON = 72.97282
GPS_RADIUS_METERS = 50

IST = pytz.timezone('Asia/Kolkata')

# --- Fingerprint retention: 1 day ---
DAYS_TO_KEEP_FINGERPRINT = 1

# --- Helper Functions ---

def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    lat1 = math.radians(lat1)
    lon1 = math.radians(lon1)
    lat2 = math.radians(lat2)
    lon2 = math.radians(lon2)
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin(dlon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def get_db_connection():
    try:
        return pymysql.connect(**DB_CONFIG)
    except Exception as e:
        print(f"❌ Database connection error: {e}")
        return None

def is_sunday(target_date):
    if isinstance(target_date, datetime):
        return target_date.weekday() == 6
    if isinstance(target_date, date):
        return target_date.weekday() == 6
    return False

# ============================================================
# AUTO-ABSENT (daily, 11:00-12:00 attendance window)
# ============================================================
def mark_absent_for_today():
    print(f"🔄 Running mark_absent_for_today at {datetime.now(IST)}")
    today = datetime.now(IST).date()

    if is_sunday(today):
        print(f"⏭️ Sunday detected, skipping auto-absent for {today}")
        return

    conn = get_db_connection()
    if not conn:
        print("❌ Database connection failed")
        return

    try:
        with conn.cursor() as cur:
            cur.execute("SELECT registration_number FROM registrations")
            all_students = [row['registration_number'] for row in cur.fetchall()]
            print(f"📋 Total students: {len(all_students)}")

            cur.execute("SELECT registration_number FROM attendance WHERE date = %s", (today,))
            marked = {row['registration_number'] for row in cur.fetchall()}
            print(f"✅ Students who marked today: {len(marked)}")

            absent_count = 0
            for reg_num in all_students:
                if reg_num not in marked:
                    cur.execute("""
                        INSERT INTO attendance
                        (registration_number, date, time_in, status, device_fingerprint)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (reg_num, today, time(0, 0), 'Absent', 'auto_absent'))
                    absent_count += 1

            conn.commit()
            print(f"✅ Marked Absent for {absent_count} students for {today}")
    except Exception as e:
        print(f"❌ Error marking absent for today: {e}")
    finally:
        conn.close()

def backfill_past_attendance():
    now = datetime.now(IST)
    print("⏳ Checking whether today's attendance should be backfilled...")
    if now.hour >= 12:
        mark_absent_for_today()
    else:
        print("⏳ Attendance window not closed yet; no backfill required.")

def cleanup_old_fingerprints():
    conn = get_db_connection()
    if not conn:
        return
    try:
        cutoff = datetime.now(IST).date() - timedelta(days=DAYS_TO_KEEP_FINGERPRINT)
        with conn.cursor() as cur:
            cur.execute("UPDATE attendance SET device_fingerprint = NULL WHERE date < %s", (cutoff,))
            updated = cur.rowcount
            conn.commit()
            print(f"🧹 Cleared fingerprints from {updated} attendance records older than {cutoff}")
    except Exception as e:
        print(f"❌ Fingerprint cleanup error: {e}")
    finally:
        conn.close()

def get_interval_status(current_time):
    hour = current_time.hour
    if 8 <= hour < 9:
        return 1
    elif 9 <= hour < 10:
        return 2
    elif 10 <= hour < 11:
        return 3
    elif 11 <= hour < 12:
        return 4
    return None

def get_month_calendar(reg_number, year, month):
    conn = get_db_connection()
    if not conn:
        return {}
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT date, status FROM attendance 
                WHERE registration_number = %s
                  AND YEAR(date) = %s
                  AND MONTH(date) = %s
                  AND DAYOFWEEK(date) <> 1
            """, (reg_number, year, month))
            records = cur.fetchall()
            day_data = {}
            for row in records:
                day = row['date'].day
                status = row['status']
                if day not in day_data:
                    day_data[day] = {'present_count': 0, 'status': 'Absent'}
                if status == 'Present':
                    day_data[day]['present_count'] += 1
                    day_data[day]['status'] = 'Present'
            return day_data
    finally:
        conn.close()

# ============================================================
# TODAY'S STATS (for dashboard blocks)
# ============================================================
def get_dashboard_stats(reg_number):
    conn = get_db_connection()
    if not conn:
        return {'present_days': 0, 'absent': 0, 'total_days': 0}
    try:
        with conn.cursor() as cur:
            today = datetime.now(IST).date()
            if is_sunday(today):
                return {'present_days': 0, 'absent': 0, 'total_days': 0}

            cur.execute("""
                SELECT COUNT(CASE WHEN status IN ('Present', 'Late') THEN 1 END) as present_days,
                       COUNT(CASE WHEN status = 'Absent' THEN 1 END) as absent
                FROM attendance 
                WHERE registration_number = %s
                  AND date = %s
                  AND DAYOFWEEK(date) <> 1
            """, (reg_number, today))
            stats = cur.fetchone()
            present_days = stats.get('present_days', 0) or 0
            absent = stats.get('absent', 0) or 0
            total_days = present_days + absent
            return {
                'present_days': present_days,
                'absent': absent,
                'total_days': total_days
            }
    finally:
        conn.close()

# ============================================================
# DAY SUMMARY (for a specific date)
# ============================================================
@app.route('/get_day_summary')
def get_day_summary():
    if not session.get('student_logged_in'):
        return jsonify({'success': False, 'message': 'Not logged in.'}), 401
    registration_number = session['registration_number']
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'success': False, 'message': 'Date parameter required'}), 400
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD'}), 400

    if is_sunday(target_date):
        return jsonify({
            'success': True,
            'present_hours': 0,
            'absent': 0,
            'total_hours': 0
        })

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed.'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(CASE WHEN status IN ('Present', 'Late') THEN 1 END) as present_hours,
                       COUNT(CASE WHEN status = 'Absent' THEN 1 END) as absent
                FROM attendance 
                WHERE registration_number = %s AND date = %s AND DAYOFWEEK(date) <> 1
            """, (registration_number, target_date))
            stats = cur.fetchone()
            present_hours = stats.get('present_hours', 0) or 0
            absent = stats.get('absent', 0) or 0
            total_hours = present_hours + absent
            return jsonify({
                'success': True,
                'present_hours': present_hours,
                'absent': absent,
                'total_hours': total_hours
            })
    finally:
        conn.close()

@app.route('/get_attendance_for_date')
def get_attendance_for_date():
    if not session.get('student_logged_in'):
        return jsonify({'success': False, 'message': 'Not logged in.'}), 401
    registration_number = session['registration_number']
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'success': False, 'message': 'Date parameter required'}), 400
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD'}), 400
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed.'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT time_in, status 
                FROM attendance 
                WHERE registration_number = %s AND date = %s
                ORDER BY id DESC
                LIMIT 1
            """, (registration_number, target_date))
            record = cur.fetchone()
            attendance_status = {}
            if record:
                time_in = record['time_in']
                if isinstance(time_in, timedelta):
                    total_seconds = int(time_in.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                elif isinstance(time_in, time):
                    time_str = time_in.strftime('%H:%M:%S')
                else:
                    time_str = str(time_in) if time_in else ''
                attendance_status = {
                    'time_in': time_str,
                    'status': record['status']
                }
            return jsonify({
                'success': True,
                'attendance_status': attendance_status,
                'marked': bool(record)
            })
    finally:
        conn.close()

# --- PUBLIC ROUTES ---
@app.route('/')
def index():
    if session.get('student_logged_in'):
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        if request.is_json:
            data = request.get_json()
            name = data.get('name', '').strip()
            registration_number = data.get('registration_number', '').strip()
            mobile = data.get('mobile', '').strip()
            year = data.get('year', '').strip()
            course = data.get('course', '').strip()
            password = data.get('password', '').strip()
            confirm_password = data.get('confirm_password', '').strip()
        else:
            name = request.form.get('name', '').strip()
            registration_number = request.form.get('registration_number', '').strip()
            mobile = request.form.get('mobile', '').strip()
            year = request.form.get('year', '').strip()
            course = request.form.get('course', '').strip()
            password = request.form.get('password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()

        year_map = {'1': 'FYBCA', '2': 'SYBCA', '3': 'TYBCA'}
        year_value = year_map.get(year, year).strip()

        if not all([name, registration_number, mobile, year_value, password, confirm_password]):
            error_msg = 'All fields are required'
            if request.is_json:
                return jsonify({'success': False, 'message': error_msg})
            return render_template('signup.html', error=error_msg)

        if year_value not in {'FYBCA', 'SYBCA', 'TYBCA'}:
            error_msg = 'Invalid year'
            if request.is_json:
                return jsonify({'success': False, 'message': error_msg})
            return render_template('signup.html', error=error_msg)

        if password != confirm_password:
            error_msg = 'Passwords do not match'
            if request.is_json:
                return jsonify({'success': False, 'message': error_msg})
            return render_template('signup.html', error=error_msg)

        if len(password) < 6:
            error_msg = 'Password must be at least 6 characters'
            if request.is_json:
                return jsonify({'success': False, 'message': error_msg})
            return render_template('signup.html', error=error_msg)

        if len(mobile) != 10 or not mobile.isdigit():
            error_msg = 'Mobile number must be 10 digits'
            if request.is_json:
                return jsonify({'success': False, 'message': error_msg})
            return render_template('signup.html', error=error_msg)

        conn = get_db_connection()
        if not conn:
            error_msg = 'Database connection failed'
            if request.is_json:
                return jsonify({'success': False, 'message': error_msg})
            return render_template('signup.html', error=error_msg)

        try:
            with conn.cursor() as cur:
                cur.execute('SELECT * FROM registrations WHERE registration_number = %s', (registration_number,))
                if cur.fetchone():
                    error_msg = 'Registration number already exists'
                    if request.is_json:
                        return jsonify({'success': False, 'message': error_msg})
                    return render_template('signup.html', error=error_msg)

                cur.execute("""
                    INSERT INTO registrations 
                    (name, registration_number, mobile, year, course, password, registered_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (name, registration_number, mobile, year_value, course, password, datetime.now(IST)))
                conn.commit()
                if request.is_json:
                    return jsonify({
                        'success': True,
                        'message': 'Account created successfully!',
                        'redirect': url_for('login')
                    })
                return redirect(url_for('login'))
        except Exception as e:
            print(f"Signup error: {e}")
            error_msg = 'An error occurred. Please try again.'
            if request.is_json:
                return jsonify({'success': False, 'message': error_msg})
            return render_template('signup.html', error=error_msg)
        finally:
            conn.close()
    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.is_json:
            data = request.get_json()
            registration_number = data.get('registration_number', '').strip()
            password = data.get('password', '').strip()
        else:
            registration_number = request.form.get('registration_number', '').strip()
            password = request.form.get('password', '').strip()

        if not registration_number or not password:
            error_msg = 'Registration number and password are required'
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({'success': False, 'message': error_msg})
            return render_template('login.html', error=error_msg)

        if registration_number == 'admin' and password == '1246':
            session['admin_logged_in'] = True
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({
                    'success': True,
                    'message': 'Admin login successful!',
                    'redirect': '/admin',
                    'is_admin': True
                })
            return redirect('/admin')

        conn = get_db_connection()
        if not conn:
            error_msg = 'Database connection failed'
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({'success': False, 'message': error_msg})
            return render_template('login.html', error=error_msg)

        try:
            with conn.cursor() as cur:
                cur.execute('SELECT * FROM registrations WHERE registration_number = %s', (registration_number,))
                user = cur.fetchone()
                if not user or user['password'] != password:
                    error_msg = 'Invalid registration number or password'
                    if request.is_json or request.headers.get('Accept') == 'application/json':
                        return jsonify({'success': False, 'message': error_msg})
                    return render_template('login.html', error=error_msg)

                session['user_id'] = user['id']
                session['registration_number'] = user['registration_number']
                session['student_name'] = user['name']
                session['student_logged_in'] = True

                if request.is_json or request.headers.get('Accept') == 'application/json':
                    return jsonify({
                        'success': True,
                        'message': 'Login successful!',
                        'redirect': url_for('dashboard')
                    })
                return redirect(url_for('dashboard'))
        except Exception as e:
            print(f"Login error: {e}")
            error_msg = 'An error occurred. Please try again.'
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({'success': False, 'message': error_msg})
            return render_template('login.html', error=error_msg)
        finally:
            conn.close()
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if not session.get('student_logged_in'):
        return redirect(url_for('login'))
    reg_number = session['registration_number']
    stats = get_dashboard_stats(reg_number)  # Today's stats
    today = datetime.now(IST).date()
    return render_template('dashboard.html',
                         student_name=session['student_name'],
                         reg_number=reg_number,
                         stats=stats,
                         current_month=today.month,
                         current_year=today.year)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/profile')
def profile():
    if not session.get('student_logged_in'):
        return redirect(url_for('login'))
    reg_number = session['registration_number']
    conn = get_db_connection()
    if not conn:
        return render_template('profile.html', student_name=session['student_name'], reg_number=reg_number, student_data={})
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM registrations WHERE registration_number = %s", (reg_number,))
            student_data = cur.fetchone()
        return render_template('profile.html',
                             student_name=session['student_name'],
                             reg_number=reg_number,
                             student_data=student_data)
    finally:
        conn.close()

@app.route('/calendar_data')
def calendar_data():
    if not session.get('student_logged_in'):
        return jsonify({'success': False, 'message': 'Not logged in.'}), 401
    year = request.args.get('year', type=int, default=datetime.now(IST).year)
    month = request.args.get('month', type=int, default=datetime.now(IST).month)
    reg_number = session['registration_number']
    day_data = get_month_calendar(reg_number, year, month)
    return jsonify({'success': True, 'data': day_data})

@app.route('/attendance_page')
def attendance_page():
    if not session.get('student_logged_in'):
        return redirect(url_for('login'))
    return render_template('attendance_form.html')

@app.route('/get_daily_attendance_status')
def get_daily_attendance_status():
    if not session.get('student_logged_in'):
        return jsonify({'success': False, 'message': 'Not logged in.'}), 401
    registration_number = session['registration_number']
    current_date = datetime.now(IST).date()
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed.'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT time_in, status
                FROM attendance
                WHERE registration_number = %s AND date = %s
                ORDER BY id DESC
                LIMIT 1
            """, (registration_number, current_date))
            record = cur.fetchone()
            attendance_status = {}
            if record:
                time_in = record['time_in']
                time_str = time_in.strftime('%H:%M:%S') if isinstance(time_in, time) else str(time_in or '')
                attendance_status = {'time_in': time_str, 'status': record['status']}
            return jsonify({'success': True, 'attendance_status': attendance_status, 'marked': bool(record)})
    finally:
        conn.close()

@app.route('/get_session_data')
def get_session_data():
    if session.get('student_logged_in'):
        return jsonify({
            'registration_number': session.get('registration_number'),
            'student_name': session.get('student_name')
        })
    return jsonify({'success': False})

# --- ATTENDANCE API (8-12, 4 intervals) ---
@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    data = request.get_json()
    registration_number = data.get('registration_number', '').strip()
    year = data.get('year', '').strip()
    latitude = data.get('latitude')
    longitude = data.get('longitude')
    device_fingerprint = data.get('device_fingerprint', '').strip()

    if registration_number.lower() == ADMIN_CODE.lower():
        session['admin_logged_in'] = True
        return jsonify({
            'success': True,
            'message': 'Admin access granted!',
            'redirect': '/admin',
            'is_admin': True
        }), 200

    if not session.get('student_logged_in'):
        return jsonify({'success': False, 'message': 'Please login first.'}), 401

    if registration_number != session['registration_number']:
        return jsonify({'success': False, 'message': 'You can only mark your own attendance.'}), 400

    if not all([registration_number, year]):
        return jsonify({'success': False, 'message': 'Registration Number and Year are required.'}), 400
    if year not in ['FYBCA', 'SYBCA', 'TYBCA']:
        return jsonify({'success': False, 'message': 'Invalid year.'}), 400
    if latitude is None or longitude is None:
        return jsonify({'success': False, 'message': 'Location access is required.'}), 400
    if not device_fingerprint:
        return jsonify({'success': False, 'message': 'Device fingerprint not available.'}), 400

    now = datetime.now(IST)
    current_time = now.time()
    current_date = now.date()

    if is_sunday(current_date):
        return jsonify({'success': False, 'message': 'Sunday is a holiday. Attendance is not counted on Sundays.'}), 400

    distance = haversine(COLLEGE_LAT, COLLEGE_LON, float(latitude), float(longitude))
    if distance > GPS_RADIUS_METERS:
        return jsonify({'success': False, 'message': f'You are not within college campus. Distance: {int(distance)} meters.'}), 400

    if current_time < time(11, 0):
        return jsonify({'success': False, 'message': 'Attendance window starts at 11:00 AM IST.'}), 400
    if current_time >= time(12, 0):
        return jsonify({'success': False, 'message': 'Attendance window closed at 12:00 PM IST.'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed.'}), 500

    try:
        with conn.cursor() as cur:
            cur.execute("SELECT year FROM registrations WHERE registration_number = %s", (registration_number,))
            student = cur.fetchone()
            if not student:
                return jsonify({'success': False, 'message': 'Invalid Registration Number.'}), 400
            if student['year'] != year:
                return jsonify({'success': False, 'message': f"Student is in {student['year']}, not {year}."}), 400

            cur.execute("""
                SELECT id, status, device_fingerprint 
                FROM attendance
                WHERE registration_number = %s AND date = %s
                LIMIT 1
            """, (registration_number, current_date))
            existing = cur.fetchone()

            if existing:
                if existing['status'] == 'Absent' and existing['device_fingerprint'] == 'auto_absent':
                    cur.execute("""
                        UPDATE attendance 
                        SET time_in = %s, status = 'Present', device_fingerprint = %s
                        WHERE id = %s
                    """, (current_time, device_fingerprint, existing['id']))
                    conn.commit()
                    return jsonify({
                        'success': True,
                        'message': 'Attendance updated to Present for today.',
                        'status': 'Present'
                    }), 200

                return jsonify({'success': False, 'message': 'Attendance already marked for today.'}), 400

            cur.execute("""
                SELECT registration_number FROM attendance
                WHERE device_fingerprint = %s AND date = %s
            """, (device_fingerprint, current_date))
            existing_device = cur.fetchone()
            if existing_device and existing_device['registration_number'] != registration_number:
                return jsonify({'success': False, 'message': 'This device has already been used for attendance today.'}), 400

            cur.execute("""
                INSERT INTO attendance
                (registration_number, date, time_in, status, device_fingerprint)
                VALUES (%s, %s, %s, %s, %s)
            """, (registration_number, current_date, current_time, 'Present', device_fingerprint))
            conn.commit()

            return jsonify({
                'success': True,
                'message': 'Attendance marked for today.',
                'status': 'Present'
            }), 200
    except Exception as e:
        print(f"❌ Mark attendance error: {e}")
        return jsonify({'success': False, 'message': 'An error occurred while marking attendance.'}), 500
    finally:
        conn.close()

# --- ADMIN ROUTES ---
@app.route('/admin_panel', methods=['GET', 'POST'])
@app.route('/admin', methods=['GET', 'POST'])
def admin_panel():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == '1246':
            session['admin_logged_in'] = True
            return redirect(url_for('admin_panel'))
        else:
            return render_template('admin.html', error="Wrong password")
    if not session.get('admin_logged_in'):
        return render_template('admin.html', error=None)
    conn = get_db_connection()
    registrations = []
    total_registrations = 0
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM registrations ORDER BY registered_at DESC")
                registrations = cur.fetchall()
                total_registrations = len(registrations)
        finally:
            conn.close()
    return render_template('admin.html', registrations=registrations, total_registrations=total_registrations)

@app.route('/admin/search')
def admin_search():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'Not authorized.'}), 401
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify({'success': True, 'data': []})
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed.'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT * FROM registrations 
                WHERE name LIKE %s OR registration_number LIKE %s
                ORDER BY registered_at DESC
            """, (f'%{query}%', f'%{query}%'))
            results = cur.fetchall()
            return jsonify({'success': True, 'data': results})
    finally:
        conn.close()

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('index'))

@app.route('/admin/reset_password', methods=['POST'])
def admin_reset_password():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    data = request.json
    reg_number = data.get('registration_number')
    new_password = data.get('new_password', '123456').strip()
    if not new_password or len(new_password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters'})
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE registrations SET password = %s WHERE registration_number = %s", (new_password, reg_number))
            conn.commit()
        return jsonify({'success': True, 'new_password': new_password})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/admin/delete_student', methods=['POST'])
def admin_delete_student():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    data = request.json
    reg_number = data.get('registration_number')
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM attendance WHERE registration_number = %s", (reg_number,))
            cur.execute("DELETE FROM registrations WHERE registration_number = %s", (reg_number,))
            conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/admin/reset_attendance', methods=['POST'])
def admin_reset_attendance():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    data = request.json
    reg_number = data.get('registration_number', '').strip()
    if not reg_number:
        return jsonify({'success': False, 'message': 'Registration number is required'}), 400
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT name FROM registrations WHERE registration_number = %s", (reg_number,))
            student = cur.fetchone()
            if not student:
                return jsonify({'success': False, 'message': 'Student not found'}), 400
            cur.execute("SELECT COUNT(*) as count FROM attendance WHERE registration_number = %s", (reg_number,))
            count_result = cur.fetchone()
            records_deleted = count_result.get('count', 0) if count_result else 0
            cur.execute("DELETE FROM attendance WHERE registration_number = %s", (reg_number,))
            conn.commit()
            return jsonify({
                'success': True,
                'message': 'Attendance records reset successfully',
                'student_name': student['name'],
                'records_deleted': records_deleted
            })
    except Exception as e:
        print(f"Error resetting attendance: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/admin/reset_all_data', methods=['POST'])
def admin_reset_all_data():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    data = request.json
    reg_number = data.get('registration_number', '').strip()
    confirm = data.get('confirm', False)
    if not reg_number:
        return jsonify({'success': False, 'message': 'Registration number is required'}), 400
    if not confirm:
        return jsonify({
            'success': False,
            'message': 'Confirmation required',
            'requires_confirmation': True
        }), 400
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT name FROM registrations WHERE registration_number = %s", (reg_number,))
            student = cur.fetchone()
            if not student:
                return jsonify({'success': False, 'message': 'Student not found'}), 400
            cur.execute("SELECT COUNT(*) as count FROM attendance WHERE registration_number = %s", (reg_number,))
            attendance_count = cur.fetchone().get('count', 0)
            cur.execute("DELETE FROM attendance WHERE registration_number = %s", (reg_number,))
            cur.execute("DELETE FROM registrations WHERE registration_number = %s", (reg_number,))
            conn.commit()
            return jsonify({
                'success': True,
                'message': 'Student and all records deleted permanently',
                'deleted_student': student['name'],
                'deleted_attendance_records': attendance_count
            })
    except Exception as e:
        print(f"Error resetting all data: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/admin/student_interval_records')
def admin_student_interval_records():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'Not authorized'}), 401
    reg_number = request.args.get('reg_number', '').strip()
    current_date = datetime.now(IST).date()
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT interval_number, time_in, status 
                FROM attendance 
                WHERE registration_number = %s AND date = %s
                ORDER BY interval_number
            """, (reg_number, current_date))
            records = cur.fetchall()
            interval_records = {}
            for record in records:
                interval = record['interval_number']
                time_in = record['time_in']
                if isinstance(time_in, timedelta):
                    total_seconds = int(time_in.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                elif isinstance(time_in, time):
                    time_str = time_in.strftime('%H:%M:%S')
                else:
                    time_str = str(time_in) if time_in else ''
                interval_records[interval] = {
                    'time_in': time_str,
                    'status': record['status']
                }
            return jsonify({
                'success': True,
                'records': interval_records
            })
    finally:
        conn.close()

# ============================================================
# DOWNLOAD FULL REPORT – 4 SHEETS (NO SLICER, NO CHART)
# ============================================================
@app.route('/download_full_report')
def download_full_report():
    if not session.get('admin_logged_in'):
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500

    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT registration_number, name, mobile, year, course, registered_at
                FROM registrations
                ORDER BY registered_at DESC
            """)
            registrations = cur.fetchall()

            cur.execute("""
                SELECT registration_number, date, time_in, status, interval_number, device_fingerprint
                FROM attendance
                WHERE DAYOFWEEK(date) <> 1
                ORDER BY date, registration_number
            """)
            attendance_records = cur.fetchall()

            cur.execute("""
                SELECT 
                    r.name,
                    r.registration_number,
                    YEAR(a.date) as year,
                    MONTH(a.date) as month,
                    WEEK(a.date, 1) as week,
                    SUM(CASE WHEN a.status = 'Present' THEN 1 ELSE 0 END) as present_count,
                    SUM(CASE WHEN a.status = 'Absent' THEN 1 ELSE 0 END) as absent_count
                FROM attendance a
                JOIN registrations r ON a.registration_number = r.registration_number
                WHERE DAYOFWEEK(a.date) <> 1
                GROUP BY r.name, r.registration_number, YEAR(a.date), MONTH(a.date), WEEK(a.date, 1)
                ORDER BY year DESC, month DESC, week, r.name
            """)
            weekly_data = cur.fetchall()

            cur.execute("""
                SELECT interval_number,
                       COUNT(CASE WHEN status = 'Present' THEN 1 END) as present_total,
                       COUNT(CASE WHEN status = 'Absent' THEN 1 END) as absent_total
                FROM attendance
                WHERE interval_number IS NOT NULL
                  AND DAYOFWEEK(date) <> 1
                GROUP BY interval_number
                ORDER BY interval_number
            """)
            interval_summary = cur.fetchall()

    except Exception as e:
        print(f"Error fetching data: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

    wb = Workbook()
    wb.remove(wb.active)

    # ============================================================
    # SHEET 1: Registrations
    # ============================================================
    ws1 = wb.create_sheet("Registrations")
    headers1 = ['Registration Number', 'Name', 'Mobile', 'Year', 'Course', 'Registered At']
    ws1.append(headers1)
    for reg in registrations:
        registered_at = reg['registered_at'].strftime('%Y-%m-%d %H:%M:%S') if reg['registered_at'] else ''
        ws1.append([
            reg['registration_number'],
            reg['name'],
            reg['mobile'],
            reg['year'],
            reg['course'] or 'BCA',
            registered_at
        ])
    for col in range(1, 7):
        ws1.column_dimensions[chr(64 + col)].width = 18

    # ============================================================
    # SHEET 2: Attendance (all records)
    # ============================================================
    ws2 = wb.create_sheet("Attendance")
    headers2 = ['Registration Number', 'Date', 'Time In', 'Status', 'Interval Number', 'Device Fingerprint']
    ws2.append(headers2)
    for rec in attendance_records:
        time_in = rec['time_in']
        if isinstance(time_in, timedelta):
            total_seconds = int(time_in.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        elif isinstance(time_in, time):
            time_str = time_in.strftime('%H:%M:%S')
        else:
            time_str = str(time_in) if time_in else ''
        ws2.append([
            rec['registration_number'],
            rec['date'].strftime('%Y-%m-%d') if rec['date'] else '',
            time_str,
            rec['status'],
            rec['interval_number'],
            rec['device_fingerprint'] or ''
        ])
    for col in range(1, 7):
        ws2.column_dimensions[chr(64 + col)].width = 18

    # ============================================================
    # SHEET 3: Student Week-wise Summary (no slicer, plain table)
    # ============================================================
    ws3 = wb.create_sheet("Student Week-wise")
    headers3 = ['Student Name', 'Registration Number', 'Year', 'Month', 'Week', 'Present', 'Absent', 'Total Days', 'Attendance %']
    ws3.append(headers3)
    for row in weekly_data:
        name = row['name']
        reg_num = row['registration_number']
        year = row['year']
        month = row['month']
        week = row['week']
        present = row['present_count']
        absent = row['absent_count']
        total = present + absent
        att_pct = round(present / total * 100, 2) if total > 0 else 0
        ws3.append([name, reg_num, year, month, week, present, absent, total, att_pct])

    # Optionally add auto-filter for convenience (but not required)
    # ws3.auto_filter.ref = ws3.dimensions

    for col in range(1, 10):
        ws3.column_dimensions[chr(64 + col)].width = 15

    # ============================================================
    # SHEET 4: Interval Analysis (data only, no chart)
    # ============================================================
    ws4 = wb.create_sheet("Interval Analysis")
    headers4 = ['Interval', 'Present Count', 'Absent Count']
    ws4.append(headers4)
    for row in interval_summary:
        ws4.append([f"Interval {row['interval_number']}", row['present_total'], row['absent_total']])

    for col in range(1, 4):
        ws4.column_dimensions[chr(64 + col)].width = 18

    # ============================================================
    # Save and send file
    # ============================================================
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'attendance_report_{datetime.now(IST).strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

# --- MANUAL TRIGGER (debugging) ---
@app.route('/test_absent/<int:interval>')
def test_absent(interval):
    if 1 <= interval <= 4:
        mark_absent_for_interval(interval)
        return f"✅ Manually triggered absent for interval {interval}"
    else:
        return "❌ Invalid interval (1-4)"

# --- SCHEDULER (4 intervals) ---
scheduler = BackgroundScheduler(timezone=IST)

scheduler.add_job(
    mark_absent_for_today,
    'cron',
    hour=12,
    minute=0,
    timezone=IST,
    id='absent_for_today'
)
print("📅 Scheduled absent for today at 12:00 IST")

scheduler.add_job(
    cleanup_old_fingerprints,
    'cron',
    hour=0,
    minute=1,
    timezone=IST,
    id='fingerprint_cleanup'
)

scheduler.start()
print("✅ Scheduler started.")
for job in scheduler.get_jobs():
    print(f"  - {job.id} | next run: {job.next_run_time}")

if __name__ == '__main__':
    backfill_past_attendance()
    app.run(debug=False, host='0.0.0.0', port=5000)